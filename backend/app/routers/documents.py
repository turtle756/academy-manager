import io
from datetime import date
from calendar import monthrange

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.auth import get_membership
from app.models.user_academy import UserAcademy
from app.models.student import Student
from app.models.classroom import Classroom, StudentClassroom
from app.models.academy import Academy
from app.models.payment import Invoice, InvoiceStatus

router = APIRouter()


def thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)

def center():
    return Alignment(horizontal="center", vertical="center")


# ============ 재원증명서 ============

@router.get("/enrollment-cert")
async def generate_enrollment_cert(
    student_id: int,
    membership: UserAcademy = Depends(get_membership),
    db: AsyncSession = Depends(get_db),
):
    """재원증명서 — 학생 1명"""
    student = await db.get(Student, student_id)
    if not student or student.academy_id != membership.academy_id:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    academy = await db.get(Academy, membership.academy_id)

    sc_result = await db.execute(
        select(StudentClassroom)
        .where(StudentClassroom.student_id == student_id)
        .options(selectinload(StudentClassroom.classroom))
    )
    classrooms = [sc.classroom.name for sc in sc_result.scalars().all() if sc.classroom]

    today_str = date.today().strftime("%Y년 %m월 %d일")
    enroll_date = student.created_at.strftime("%Y년 %m월 %d일") if student.created_at else "-"

    wb = Workbook()
    ws = wb.active
    ws.title = "재원증명서"
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 32
    ws.row_dimensions[1].height = 20

    # 제목
    ws.merge_cells('A1:B1')
    ws['A1'] = '재  원  증  명  서'
    ws['A1'].font = Font(size=18, bold=True)
    ws['A1'].alignment = center()

    ws.append(['', ''])
    ws.append(['발급일', today_str])
    ws.append(['', ''])

    # 학원 정보
    ws.append(['학원명', academy.name if academy else ''])
    ws.append(['주소', (academy.address or '') + ' ' + (academy.address_detail or '') if academy else ''])
    ws.append(['연락처', academy.phone or '' if academy else ''])
    ws.append(['', ''])

    # 학생 정보
    ws.append(['성명', student.name])
    ws.append(['학교 / 학년', f"{student.school or ''} / {student.grade or ''}"])
    ws.append(['수강반', ', '.join(classrooms) if classrooms else '-'])
    ws.append(['등록일', enroll_date])
    ws.append(['', ''])

    ws.merge_cells(f'A{ws.max_row + 1}:B{ws.max_row + 1}')
    last_row = ws.max_row + 1
    ws.merge_cells(f'A{last_row}:B{last_row}')
    ws.cell(last_row, 1, f'위 학생이 본 학원에 재원 중임을 증명합니다.')
    ws.cell(last_row, 1).alignment = center()
    ws.append(['', ''])
    ws.append(['', today_str])
    ws.append(['', f'{academy.name if academy else ""} 원장 (인)'])

    # 테두리
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row - 3):
        for cell in row:
            cell.border = thin_border()

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=재원증명서_{student.name}.xlsx"},
    )


# ============ 납부확인서 ============

@router.get("/payment-cert")
async def generate_payment_cert(
    student_id: int,
    month: str,  # "2026-04"
    membership: UserAcademy = Depends(get_membership),
    db: AsyncSession = Depends(get_db),
):
    """납부확인서 — 학생 + 월"""
    student = await db.get(Student, student_id)
    if not student or student.academy_id != membership.academy_id:
        raise HTTPException(status_code=404, detail="학생을 찾을 수 없습니다")

    academy = await db.get(Academy, membership.academy_id)

    year, m = month.split("-")
    year_i, m_i = int(year), int(m)
    _, num_days = monthrange(year_i, m_i)
    start = date(year_i, m_i, 1)
    end = date(year_i, m_i, num_days)

    inv_result = await db.execute(
        select(Invoice).where(
            Invoice.student_id == student_id,
            Invoice.academy_id == membership.academy_id,
            Invoice.due_date >= start,
            Invoice.due_date <= end,
        )
    )
    invoices = inv_result.scalars().all()

    today_str = date.today().strftime("%Y년 %m월 %d일")
    month_label = f"{year_i}년 {m_i}월"

    wb = Workbook()
    ws = wb.active
    ws.title = "납부확인서"
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 32

    ws.merge_cells('A1:B1')
    ws['A1'] = '납  부  확  인  서'
    ws['A1'].font = Font(size=18, bold=True)
    ws['A1'].alignment = center()

    ws.append(['', ''])
    ws.append(['발급일', today_str])
    ws.append(['', ''])
    ws.append(['학원명', academy.name if academy else ''])
    ws.append(['연락처', academy.phone or '' if academy else ''])
    ws.append(['', ''])
    ws.append(['성명', student.name])
    ws.append(['납부 월', month_label])
    ws.append(['', ''])

    # 납부 내역
    ws.merge_cells(f'A{ws.max_row + 1}:B{ws.max_row + 1}')
    header_row = ws.max_row + 1
    ws.merge_cells(f'A{header_row}:B{header_row}')
    ws.cell(header_row, 1, '< 납부 내역 >')
    ws.cell(header_row, 1).font = Font(bold=True)
    ws.cell(header_row, 1).alignment = center()

    ws.append(['항목', '금액'])
    ws[f'A{ws.max_row}'].font = Font(bold=True)
    ws[f'B{ws.max_row}'].font = Font(bold=True)

    total = 0
    for inv in invoices:
        status_label = '납부완료' if inv.status == InvoiceStatus.PAID else '미납'
        desc = inv.description or '수강료'
        ws.append([f"{desc} ({status_label})", f"{inv.amount:,}원"])
        if inv.status == InvoiceStatus.PAID:
            total += inv.amount

    ws.append(['', ''])
    ws.append(['합계 (납부 완료)', f"{total:,}원"])
    ws[f'A{ws.max_row}'].font = Font(bold=True)
    ws[f'B{ws.max_row}'].font = Font(bold=True)

    ws.append(['', ''])
    cert_row = ws.max_row + 1
    ws.merge_cells(f'A{cert_row}:B{cert_row}')
    ws.cell(cert_row, 1, f'위 금액을 납부하였음을 확인합니다.')
    ws.cell(cert_row, 1).alignment = center()

    ws.append(['', ''])
    ws.append(['', today_str])
    ws.append(['', f'{academy.name if academy else ""} 원장 (인)'])

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row - 3):
        for cell in row:
            cell.border = thin_border()

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=납부확인서_{student.name}_{month}.xlsx"},
    )
